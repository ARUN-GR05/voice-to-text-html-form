"""
Flask Application for Audio Transcription

This is the main web application that serves the HTML form and handles
audio uploads. It uses the OpenAI Whisper API to transcribe the uploaded
audio and returns the result.
"""
from flask import Flask, request, jsonify, render_template
import tempfile
import os
from openai import OpenAI
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv
import imageio_ffmpeg
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

from io import BytesIO

load_dotenv()

# Configure pydub to use imageio-ffmpeg BEFORE importing pydub
from pydub import AudioSegment
AudioSegment.converter = imageio_ffmpeg.get_ffmpeg_exe()
AudioSegment.ffmpeg = imageio_ffmpeg.get_ffmpeg_exe()
AudioSegment.ffprobe = imageio_ffmpeg.get_ffmpeg_exe()

app = Flask(__name__)

# --- Configuration ---
# REPLACE THESE WITH YOUR ACTUAL VALUES OR SET ENVIRONMENT VARIABLES
API_KEY = os.environ.get("OPENAI_API_KEY", "YOUR_API_KEY_HERE")
BASE_URL = os.environ.get("OPENAI_BASE_URL", "YOUR_BASE_URL_HERE") 

client = OpenAI(
    api_key=API_KEY,
    base_url=BASE_URL
)

def convert_to_wav(input_path):
    audio = AudioSegment.from_file(input_path)
    wav_path = input_path + ".wav"
    audio.export(wav_path, format="wav")
    return wav_path

import csv

# ... (rest of imports and setup)

def transcribe_audio(audio_path):
    """
    Transcribes audio using the configured OpenAI-compatible API.
    Returns only the transcribed text, stripping any metadata.
    """
    with open(audio_path, "rb") as audio_file:
        transcript = client.audio.transcriptions.create(
            model="whisper-1", 
            file=audio_file,
            response_format="text"
        )
    
def transcribe_audio(audio_path):
    """
    Transcribes audio using the configured OpenAI-compatible API.
    Returns only the transcribed text, stripping any metadata.
    """
    try:
        with open(audio_path, "rb") as audio_file:
            transcript = client.audio.transcriptions.create(
                model="whisper-1", 
                file=audio_file,
                response_format="text"
            )
        
        # Robust handling for various API response formats
        if isinstance(transcript, dict):
            return transcript.get("text", "").strip()
        
        if hasattr(transcript, 'text'):
            text_val = transcript.text.strip()
            import json
            try:
                parsed = json.loads(text_val)
                if isinstance(parsed, dict) and "text" in parsed:
                    return parsed["text"].strip()
            except:
                pass
            return text_val
        
        if isinstance(transcript, str):
            text_val = transcript.strip()
            import json
            try:
                parsed = json.loads(text_val)
                if isinstance(parsed, dict) and "text" in parsed:
                    return parsed["text"].strip()
            except:
                pass
            return text_val
            
        return str(transcript).strip()
    except Exception as e:
        print(f"Transcription error: {e}")
        raise e

def ai_correct_text(text):
    """
    Uses LLM to correct medical terminology, grammar, and structure.
    """
    if not text or len(text) < 3:
        return text
    
    try:
        response = client.chat.completions.create(
            model="gpt-3.5-turbo", # Or your preferred model
            messages=[
                {"role": "system", "content": "You are a professional medical scribe. Correct the following medical transcript for grammar, medical spelling, and professional structure. Maintain the original meaning exactly. If the input is just a few words, return them as is but correctly spelled."},
                {"role": "user", "content": text}
            ],
            temperature=0.3
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        print(f"AI Correction error: {e}")
        return text

@app.route('/')
def index():
    return render_template('form.html')

@app.route('/transcribe', methods=['POST'])
def transcribe():
    if 'audio' not in request.files:
        return jsonify({'error': 'No audio file provided'}), 400

    file = request.files['audio']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400

    with tempfile.NamedTemporaryFile(suffix=os.path.splitext(file.filename)[1], delete=False) as temp_audio_file:
        temp_audio_file.write(file.read())
        temp_audio_file_path = temp_audio_file.name

    try:
        if not temp_audio_file_path.lower().endswith('.wav'):
             wav_path = convert_to_wav(temp_audio_file_path)
             try: os.remove(temp_audio_file_path)
             except: pass
        else:
             wav_path = temp_audio_file_path

    except Exception as e:
        if os.path.exists(temp_audio_file_path):
            os.remove(temp_audio_file_path)
        return jsonify({'error': 'Audio conversion failed: ' + str(e)}), 500

    try:
        # 1. Transcribe audio
        raw_text = transcribe_audio(wav_path)
        
        # 2. AI Correction (Optional, can be toggled via frontend in future)
        # For now, let's provide it as part of the result
        corrected_text = ai_correct_text(raw_text)
        
    except Exception as e:
        return jsonify({'error': 'Processing failed: ' + str(e)}), 500
    finally:
        if os.path.exists(wav_path):
            os.remove(wav_path)

    return jsonify({
        'raw_transcription': raw_text,
        'paraphrased_text': corrected_text 
    })

@app.route('/submit', methods=['POST'])
def submit():
    data = request.json
    excel_file = 'data.xlsx'
    csv_file = 'data.csv'
    headers = [
        "Name", "Place", "Age", "Gender", "Symptoms", "Diagnosis", "Prescription"
    ]
    row_data = [
        data.get('name'),
        data.get('place'),
        data.get('age'),
        data.get('gender'),
        data.get('symptoms'),
        data.get('diagnosis'),
        data.get('prescription')
    ]

    # --- Save to Excel ---
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
    else:
        wb = load_workbook(excel_file)
        ws = wb.active
    ws.append(row_data)
    wb.save(excel_file)

    # --- Save to CSV ---
    file_exists = os.path.isfile(csv_file)
    with open(csv_file, mode='a', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(headers)
        writer.writerow(row_data)

    return jsonify({'message': 'Data saved successfully to Excel and CSV.'})

@app.route('/generate_pdf', methods=['POST'])
def generate_pdf():
    if not HAS_REPORTLAB:
        return jsonify({'error': 'PDF generation library (reportlab) is not installed on this system.'}), 501
    
    data = request.json
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # --- Header ---
    p.setFillColor(colors.HexColor("#0f172a")) # Primary Navy
    p.rect(0, height-80, width, 80, fill=1)
    
    p.setFillColor(colors.white)
    p.setFont("Helvetica-Bold", 24)
    p.drawString(40, height-50, "MEDICAL PRESCRIPTION")
    
    p.setFont("Helvetica", 10)
    p.drawString(width-180, height-30, "ClinicaSync Digital Scribe")
    p.drawString(width-180, height-45, f"Date: {data.get('date', 'Today')}")

    # --- Patient Info ---
    p.setFillColor(colors.black)
    p.setFont("Helvetica-Bold", 14)
    p.drawString(40, height-120, "PATIENT INFORMATION")
    p.line(40, height-125, 200, height-125)
    
    p.setFont("Helvetica", 12)
    p.drawString(40, height-150, f"Name: {data.get('name', 'N/A')}")
    p.drawString(40, height-170, f"Age: {data.get('age', 'N/A')}")
    p.drawString(250, height-150, f"Gender: {data.get('gender', 'N/A')}")
    p.drawString(250, height-170, f"Place: {data.get('place', 'N/A')}")

    # --- Clinical Details ---
    y = height - 220
    sections = [
        ("SYMPTOMS", data.get('symptoms', 'N/A')),
        ("DIAGNOSIS", data.get('diagnosis', 'N/A')),
        ("PRESCRIPTION", data.get('prescription', 'N/A'))
    ]

    for title, content in sections:
        p.setFont("Helvetica-Bold", 12)
        p.setFillColor(colors.HexColor("#0d9488")) # Secondary Teal
        p.drawString(40, y, title)
        y -= 20
        p.setFillColor(colors.black)
        p.setFont("Helvetica", 11)
        
        text_obj = p.beginText(40, y)
        text_obj.setTextOrigin(40, y)
        text_obj.setFont("Helvetica", 11)
        
        words = content.split()
        line = ""
        for word in words:
            if p.stringWidth(line + " " + word) < width - 80:
                line += " " + word
            else:
                text_obj.textLine(line.strip())
                line = word
                y -= 15
        text_obj.textLine(line.strip())
        p.drawText(text_obj)
        y -= 40

    p.setFont("Helvetica-Oblique", 9)
    p.setFillColor(colors.gray)
    p.drawCentredString(width/2, 40, "Digitally generated by ClinicaSync AI Medical Scribe")

    p.showPage()
    p.save()
    
    buffer.seek(0)
    from flask import send_file
    return send_file(buffer, as_attachment=True, download_name=f"prescription_{data.get('name', 'patient')}.pdf", mimetype='application/pdf')

@app.route('/dashboard')
def dashboard():
    entries = []
    csv_file = 'data.csv'
    if os.path.exists(csv_file):
        with open(csv_file, mode='r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            entries = list(reader)
    return render_template('dashboard.html', entries=entries)

if __name__ == '__main__':
    app.run(debug=True)
